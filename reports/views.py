from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, TemplateView, View
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404
from django.urls import reverse_lazy
from django.utils import timezone
from django.contrib import messages
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta

from organisations.middleware import TenantContextMixin
from reports.services import (
    TrialBalanceService, BalanceSheetService, IncomeStatementService,
    CashFlowService, AgedReceivablesService, AgedPayablesService
)


class ReportBaseView(LoginRequiredMixin, TenantContextMixin, TemplateView):
    """Base view for reports."""

    def get_date_params(self):
        today = timezone.now().date()
        as_of = self.request.GET.get('as_of')
        if as_of:
            as_of = date.fromisoformat(as_of)
        else:
            as_of = today

        start_date = self.request.GET.get('start_date')
        if start_date:
            start_date = date.fromisoformat(start_date)
        else:
            start_date = as_of.replace(day=1)

        end_date = self.request.GET.get('end_date')
        if end_date:
            end_date = date.fromisoformat(end_date)
        else:
            end_date = as_of

        period_type = self.request.GET.get('period_type', 'monthly')

        return {
            'as_of': as_of,
            'start_date': start_date,
            'end_date': end_date,
            'period_type': period_type,
        }


class FinancialReportsIndexView(LoginRequiredMixin, TenantContextMixin, TemplateView):
    template_name = 'reports/index.html'


class TrialBalanceReportView(ReportBaseView):
    template_name = 'reports/trial_balance.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = self.get_date_params()

        context['report'] = TrialBalanceService.generate(
            organisation=self.request.tenant,
            as_of=params['as_of'],
        )
        context['as_of'] = params['as_of']
        return context


class BalanceSheetReportView(ReportBaseView):
    template_name = 'reports/balance_sheet.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = self.get_date_params()

        context['report'] = BalanceSheetService.generate(
            organisation=self.request.tenant,
            as_of=params['as_of'],
        )
        context['as_of'] = params['as_of']
        return context


class IncomeStatementReportView(ReportBaseView):
    template_name = 'reports/income_statement.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = self.get_date_params()

        context['report'] = IncomeStatementService.generate(
            organisation=self.request.tenant,
            start_date=params['start_date'],
            end_date=params['end_date'],
            include_prev_period=True,
            period_type=params['period_type'],
        )
        context['start_date'] = params['start_date']
        context['end_date'] = params['end_date']
        return context


class CashFlowReportView(ReportBaseView):
    template_name = 'reports/cash_flow.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = self.get_date_params()

        context['report'] = CashFlowService.generate(
            organisation=self.request.tenant,
            start_date=params['start_date'],
            end_date=params['end_date'],
        )
        context['start_date'] = params['start_date']
        context['end_date'] = params['end_date']
        return context


class AgedReceivablesReportView(ReportBaseView):
    template_name = 'reports/aged_receivables.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = self.get_date_params()

        context['report'] = AgedReceivablesService.generate(
            organisation=self.request.tenant,
            as_of=params['as_of'],
        )
        context['as_of'] = params['as_of']
        return context


class AgedPayablesReportView(ReportBaseView):
    template_name = 'reports/aged_payables.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = self.get_date_params()

        context['report'] = AgedPayablesService.generate(
            organisation=self.request.tenant,
            as_of=params['as_of'],
        )
        context['as_of'] = params['as_of']
        return context


class GeneralLedgerReportView(ReportBaseView):
    template_name = 'reports/general_ledger.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = self.get_date_params()

        from ledger.models import Account, JournalEntryLine

        accounts = Account.objects.filter(
            organisation=self.request.tenant,
            allow_transactions=True,
        ).order_by('code')

        account_id = self.request.GET.get('account')
        if account_id:
            accounts = accounts.filter(id=account_id)

        ledger_data = []
        for account in accounts:
            lines = JournalEntryLine.objects.filter(
                account=account,
                entry__status='posted',
                entry__date__gte=params['start_date'],
                entry__date__lte=params['end_date'],
            ).select_related('entry').order_by('entry__date')

            if lines.exists():
                running_balance = account.opening_balance
                entries = []
                for line in lines:
                    if line.debit_amount > 0:
                        running_balance += line.debit_amount
                    else:
                        running_balance -= line.credit_amount

                    entries.append({
                        'date': line.entry.date,
                        'entry_number': line.entry.entry_number,
                        'description': line.description or line.entry.description,
                        'debit': line.debit_amount,
                        'credit': line.credit_amount,
                        'balance': running_balance,
                    })

                ledger_data.append({
                    'account': account,
                    'entries': entries,
                    'beginning_balance': account.opening_balance,
                    'ending_balance': running_balance,
                })

        context['ledger_data'] = ledger_data
        context['accounts'] = Account.objects.filter(organisation=self.request.tenant).order_by('code')
        context['start_date'] = params['start_date']
        context['end_date'] = params['end_date']
        return context


class ReportExportView(LoginRequiredMixin, TenantContextMixin, View):
    """Export report to PDF or Excel."""

    def get(self, request, report_type):
        format_type = request.GET.get('format', 'pdf')

        # Get date params
        params = {
            'as_of': date.fromisoformat(request.GET.get('as_of', timezone.now().date().isoformat())),
            'start_date': date.fromisoformat(request.GET.get('start_date', timezone.now().date().replace(day=1).isoformat())),
            'end_date': date.fromisoformat(request.GET.get('end_date', timezone.now().date().isoformat())),
        }

        # Generate report data
        if report_type == 'trial-balance':
            report = TrialBalanceService.generate(request.tenant, params['as_of'])
            report_name = f"Trial_Balance_{params['as_of']}"
        elif report_type == 'balance-sheet':
            report = BalanceSheetService.generate(request.tenant, params['as_of'])
            report_name = f"Balance_Sheet_{params['as_of']}"
        elif report_type == 'income-statement':
            report = IncomeStatementService.generate(request.tenant, params['start_date'], params['end_date'])
            report_name = f"Income_Statement_{params['start_date']}_to_{params['end_date']}"
        else:
            return JsonResponse({'error': 'Invalid report type'}, status=400)

        if format_type == 'pdf':
            return self._export_pdf(report_name, report, report_type)
        elif format_type == 'excel':
            return self._export_excel(report_name, report, report_type)
        else:
            return JsonResponse({'error': 'Invalid format'}, status=400)

    def _export_pdf(self, filename, report, report_type):
        """Generate PDF export."""
        from reportlab.pdfgen import canvas
        from io import BytesIO

        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        p.setFont('Helvetica-Bold', 16)
        p.drawString(100, 750, f"Accounting SaaS - {report.get('organisation').name}")

        p.setFont('Helvetica-Bold', 12)
        p.drawString(100, 720, f"Report: {report_type.replace('-', ' ').title()}")

        y = 680
        p.setFont('Helvetica', 10)

        if report_type == 'income-statement':
            p.drawString(100, y, f"Period: {report['start_date']} to {report['end_date']}")
            y -= 30

            p.drawString(100, y, "Revenue")
            y -= 20
            for item in report['revenue']:
                p.drawString(120, y, f"{item['account'].code} - {item['account'].name}: ${item['current']:,.2f}")
                y -= 15

            p.drawString(100, y, f"Total Revenue: ${report['total_revenue']:,.2f}")
            y -= 30

            p.drawString(100, y, "Expenses")
            y -= 20
            for item in report['expenses']:
                p.drawString(120, y, f"{item['account'].code} - {item['account'].name}: ${item['current']:,.2f}")
                y -= 15

            p.drawString(100, y, f"Total Expenses: ${report['total_expenses']:,.2f}")
            y -= 30

            p.drawString(100, y, f"Net Income: ${report['net_income']:,.2f}")

        p.showPage()
        p.save()

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response

    def _export_excel(self, filename, report, report_type):
        """Generate Excel export."""
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.replace('-', ' ').title()

        # Header styles
        bold_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # Title
        ws['A1'] = f"Accounting SaaS - {report.get('organisation').name}"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A2'] = f"Report: {report_type.replace('-', ' ').title()}"
        ws['A2'].font = bold_font

        if report_type == 'income-statement':
            ws['A3'] = f"Period: {report['start_date']} to {report['end_date']}"

            row = 5
            ws[f'A{row}'] = 'Revenue'
            ws[f'A{row}'].font = bold_font

            row += 1
            for item in report['revenue']:
                ws[f'A{row}'] = f"{item['account'].code} - {item['account'].name}"
                ws[f'C{row}'] = item['current']
                row += 1

            ws[f'A{row}'] = 'Total Revenue'
            ws[f'A{row}'].font = bold_font
            ws[f'C{row}'] = report['total_revenue']

            row += 2
            ws[f'A{row}'] = 'Expenses'
            ws[f'A{row}'].font = bold_font

            row += 1
            for item in report['expenses']:
                ws[f'A{row}'] = f"{item['account'].code} - {item['account'].name}"
                ws[f'C{row}'] = item['current']
                row += 1

            ws[f'A{row}'] = 'Total Expenses'
            ws[f'A{row}'].font = bold_font
            ws[f'C{row}'] = report['total_expenses']

            row += 2
            ws[f'A{row}'] = 'Net Income'
            ws[f'A{row}'].font = bold_font
            ws[f'C{row}'] = report['net_income']

        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        response = HttpResponse(
            openpyxl.utils.get_response(wb),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response
